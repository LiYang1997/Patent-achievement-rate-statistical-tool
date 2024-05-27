import sys
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


def char_width(char):
    """根据字符类型返回预估宽度。"""
    if '\u4e00' <= char <= '\u9fff':
        return 2  # 中文字符
    elif char.isdigit():
        return 1.5  # 数字
    elif char.isalpha():
        return 1  # 英文字符
    else:
        return 1.2  # 符号和其他字符

def get_adjusted_width(names):
    """计算整列内容的最大预估宽度。"""
    max_length = 0
    for name in names:
        total_width = sum(char_width(char) for char in name)
        max_length = max(max_length, total_width)
    return max_length

def create_excel(names, team_objectives, filename='专利达成情况.xlsx'):
    """创建Excel文件并填充内容。"""
    # 按首字母排序
    sorted_names = sorted(names, key=lambda name: name[0])

    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active

    # 合并第一行和第二行的20列
    ws.merge_cells('A1:T2')

    # 计算个人目标
    total_names = len(sorted_names)
    individual_objective = team_objectives / total_names if total_names > 0 else 0

    # 设置合并后的单元格的内容
    ws['A1'] = f'团队目标：{team_objectives} 个人目标：{individual_objective}'

#     # 设置所有单元格自动换行
#     for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
#         for cell in row:
#             print(ws(row).)
#             cell.alignment = Alignment(wrap_text=True)
    # 设置第四行的内容
    ws['A4'] = '个人达成率0.00%'

    # 将排序后的人名填充到第五行及以下
    for row, name in enumerate(sorted_names, start=5):
        cell = f'A{row}'
        ws[cell] = name

    # 调整列宽
    adjusted_width = get_adjusted_width(['个人达成率0.00%'] + sorted_names)
    if adjusted_width > 0:
        # 设置列宽
        ws.column_dimensions[get_column_letter(1)].width = adjusted_width

    # 设置文本居中
#     for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=20):
#         for cell in row:
#             cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

    # 保存工作簿为指定的文件名
    wb.save(filename)

def calculate_individual_achievement(total_names, old_achievement, completed_count, team_objectives):
    """计算个人达成率。"""
    return old_achievement + (completed_count / team_objectives) if total_names > 0 else 0

def sort_columns_by_achievement(ws):
    # 交换两列的函数
    def swap_columns(ws, col1, col2):
        # 首先确定所有影响到的合并单元格的最大行号
        max_merged_row = 1  # 默认从第一行开始，如果没有合并单元格，则从第一行开始交换
        for merge_range in ws.merged_cells.ranges:
            if merge_range.min_col <= max(col1, col2) and merge_range.max_col >= min(col1, col2):
                # 更新最大行号
                max_merged_row = max(max_merged_row, merge_range.max_row)

        # 从合并单元格下面的行开始交换列数据
        for row in ws.iter_rows(min_row=max_merged_row + 1, max_row=ws.max_row, min_col=min(col1, col2), max_col=max(col2, col1)):
            # 计算列索引偏移
            index1 = col1 - min(col1, col2)
            index2 = col2 - min(col1, col2)
            # 交换数据
            row[index1].value, row[index2].value = row[index2].value, row[index1].value
#     def swap_columns(ws, col1, col2):
#         for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=min(col1, col2), max_col=max(col2, col1)):
#             row[col2 - min(col1, col2)].value, row[col1 - min(col1, col2)].value = \
#             row[col1 - min(col1, col2)].value, row[col2 - min(col1, col2)].value

    # 遍历所有列，对比并交换列
#     print(ws.cell(row=4, column=1).value)
    for i in range(1, ws.max_column):
        for j in range(i + 1, ws.max_column):
            # 获取列中的个人达成率值
            cell_value_i = ws.cell(row=4, column=i).value
            cell_value_j = ws.cell(row=4, column=j).value
            
            if '个人达成率' in str(cell_value_i) and '个人达成率' in str(cell_value_j):
                rate_i = float(cell_value_i.split('个人达成率')[1].replace('%', '').strip())
                rate_j = float(cell_value_j.split('个人达成率')[1].replace('%', '').strip())
                # print(cell_value_i,cell_value_j,i,j)
                # print(rate_i,rate_j)
                # 如果当前列的个人达成率小于相邻列，则交换两列
                if rate_i < rate_j:
                    swap_columns(ws, i, j)

def new_idea(filename, name, completed_count):
    """处理新的主意并更新Excel文件。"""
    try:
        wb = load_workbook(filename)
        ws = wb.active

        # 从第一行获取个人目标值
        individual_objective = float(ws['A1'].value.split('个人目标：')[1].strip())

        # 查找名字所在的列
        name_col_index = None
        for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=5, max_row=ws.max_row):
            for tmp in col:
                if tmp.value == name:
                    name_col_index = tmp.column
                #     print('name_col_index')
                #     print(name_col_index)
                    break

        # 在名字所在的列中查找个人达成率的行
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=name_col_index, max_col=name_col_index):
        #     if row[0].value == '个人达成率0%':
            if '个人达成率' in str(row[0].value):
                # 提取百分比数值部分，并去除末尾的百分号和空格
                old_achievement_str = row[0].value.split('个人达成率')[1].replace('%', '').strip()
                # 将提取的字符串转换为浮点数
                old_achievement = float(old_achievement_str) / 100  # 转换为小数形式
                # print('old_achievement=',old_achievement)
                achievement_row_index = row[0].row
                # print('achievement_row_index')
                # print(achievement_row_index)
            if row[0].value == name:
                name_row_index = row[0].row
                row[0].value =None#删除原来的名字
                # print('name_row_index')
                # print(name_row_index)
                break

        if name_col_index and achievement_row_index:
            # 计算个人达成率
            total_names = ws.max_row - 4
        #     print('total_names')
        #     print(total_names)
            individual_achievement = calculate_individual_achievement(total_names, old_achievement, completed_count, individual_objective)
        
            not_insert = 0
            for row in ws.iter_rows(min_row=achievement_row_index, max_row=achievement_row_index, min_col=1, max_col=ws.max_column):
                for tmp in row:
                    if '个人达成率' in str(tmp.value):
                        print(tmp.value)
                        # 提取百分比数值部分，并去除末尾的百分号和空格
                        old_achievement_str = tmp.value.split('个人达成率')[1].replace('%', '').strip()
                        # 将提取的字符串转换为浮点数
                        old_achievement = float(old_achievement_str) / 100  # 转换为小数形式
                        print('individual_achievement',individual_achievement)
                        print('old_achievement',old_achievement)
                        if individual_achievement == old_achievement:
                            not_insert = 1
                            find_name_col_index = tmp.column
                            break
                # 插入新的个人达成率列
            if not not_insert:
                tmp = ws['A1'].value
                ws.insert_cols(name_col_index)
                new_col_letter = get_column_letter(name_col_index)
                ws[f'{new_col_letter}4'] = f'个人达成率{individual_achievement:.2%}'
                ws['A1'].value = tmp
                find_name_col_index = name_col_index
                collating_col_index = name_col_index + 1
            else:
                new_col_letter = get_column_letter(find_name_col_index)
                collating_col_index = name_col_index

            # 将名字移动到个人达成率x下面第一个非空的单元格
            # 首先找到个人达成率x所在的行
            achievement_x_row_index = None
            for row in ws.iter_rows(min_row=achievement_row_index + 1, max_row=ws.max_row, min_col=find_name_col_index, max_col=find_name_col_index):
                if row[0].value is None:
                    achievement_x_row_index = row[0].row
                    print(achievement_x_row_index)
                    break

            # 如果找到，则将名字移动到该位置
            if achievement_x_row_index:
                ws[f'{new_col_letter}{achievement_x_row_index}'] = name
                # ws[f'{new_col_letter}{achievement_x_row_index}'].alignment = Alignment(horizontal='center', vertical='center')
                # 调整列宽
                # adjusted_width = 20#get_adjusted_width(['个人达成率0%'] + sorted_names)
                # if adjusted_width > 0:
                # # 设置列宽
                #     ws.column_dimensions[get_column_letter(1)].width = adjusted_width
            # 检查个人达成率xxx下面是否有名字信息
            has_names_below_achievement = False
            for row in ws.iter_rows(min_row=achievement_row_index + 1, max_row=ws.max_row, min_col=collating_col_index, max_col=collating_col_index):
                if row[0].value is not None:
                    has_names_below_achievement = True
                    break

            # 如果没有名字信息，则删除整列
            if not has_names_below_achievement:
                tmp = ws['A1'].value
                ws.delete_cols(collating_col_index)
                ws['A1'].value = tmp
            else:
                # 覆盖旧列空的单元格
                max_row = ws.max_row
                for row in range(achievement_row_index + 1,max_row, 1):
                    current_cell = ws.cell(row=row, column=collating_col_index)
                    next_cell = ws.cell(row=row + 1, column=collating_col_index)
                    # print(current_cell.value)
                    # print(next_cell.value)
                    if current_cell.value is None and next_cell.value is not None:
                        current_cell.value = next_cell.value
                        next_cell.value = None
                        
            sort_columns_by_achievement(ws)
            print(f"成功更新 {name} 的个人达成率为 {individual_achievement:.2%}")
            for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1, max_row=ws.max_row):
                column_has_achievement = False
                column_data_below_achievement = []

                for cell in col:
                    if '个人达成率' in str(cell.value):
                        column_has_achievement = True
                #     elif column_has_achievement and cell.value:
                        column_data_below_achievement.append(str(cell.value))
                        break

                if column_has_achievement and column_data_below_achievement:
                    adjusted_width = get_adjusted_width(column_data_below_achievement)
                    ws.column_dimensions[get_column_letter(cell.column)].width = adjusted_width
                #     print('adjusted_width',adjusted_width)

            # 保存工作簿
            wb.save(filename)
            
        else:
            print(f"未找到名字 {name} 或 '个人达成率' 列，请确认名字是否正确或列是否存在。")
    except Exception as e:
        print(f"处理主意时出错：{e}")
        
def main():
    """主函数，处理命令行参数并调用create_excel函数。"""
    # 解析命令行参数
    args = sys.argv[1:]
    filename = ''


    # 参数-c后面跟着文件名
    if '-c' in args:
        c_index = args.index('-c')
        filename = args[c_index + 1]
        # 参数-names后面跟着名字列表
        if '-names' in args:
            names_index = args.index('-names')
            names = args[names_index + 1:]
            # 参数-team-objectives后面跟着团队目标
            if '-team-objectives' in args:
                team_objectives_index = args.index('-team-objectives')
                team_objectives = int(args[team_objectives_index + 1])

                # 调用create_excel函数
                create_excel(names, team_objectives, filename)
    if '-f' in args:
        f_index = args.index('-f')
        filename = args[f_index + 1]
        # wb = load_workbook(filename)
        # ws = wb.active
        # print('test\n')
        # sort_columns_by_achievement(ws)
        # wb.save(filename)
        if '-idea' in args:
            idea_index = args.index('-idea')
            name = args[idea_index + 1]
            completed_count = float(args[idea_index + 2])

            # 调用new_idea函数
            new_idea(filename, name, completed_count)


if __name__ == '__main__':
    main()
